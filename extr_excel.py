from openpyxl import load_workbook
from openpyxl.styles import Font

def format_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Форматирование заголовков
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        logger.info(f"Excel-файл '{excel_path}' успешно отформатирован.")
    except Exception as e:
        logger.error(f"Ошибка при форматировании Excel-файла: {e}")
        logger.exception("Трассировка ошибки:")

if __name__ == "__main__":
    db_path = "emails.db"  # Укажите путь к вашей базе данных
    excel_path = "extracted_data.xlsx"  # Укажите путь для сохранения Excel-файла
    
    # Извлечение данных
    extracted_data = extract_price_and_route(db_path)
    
    # Отображение данных в консоли
    display_extracted_data(extracted_data)
    
    # Экспорт данных в Excel
    export_to_excel(extracted_data, excel_path)
    
    # Форматирование Excel-файла
    format_excel(excel_path)
